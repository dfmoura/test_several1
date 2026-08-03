#!/usr/bin/env python3
"""
Sincroniza catálogo + mapa de facas a partir de ORcAMENTO_OFICIAL_*.xlsm.

Fonte do mapa: sheet 'MAPA DE FACAS 20260715 ATUAL'
(A aba MAPA_DE_FACAS é só o pivot/seleção do job → ORÇAMENTO.)

Regras de medida (planilha oficial, com muito preenchimento manual):
  - REDONDA: TAMANHO = diâmetro (Ø). Medida exibida = Ø {diametro}.
  - TAMANHO com "X" (OVAL, ESPECIAL, DESENHADA, LACRE…): medida = TAMANHO.
  - RETA (altura numérica): medida = LARGURA x TAMANHO (quando houver largura).
  - PUXADA: alimenta o cálculo (D8). Pode estar vazia → dados incompletos.
  - REP: REPETIÇÃO (coluna REP).

Uso:
  python backend/scripts/sync_from_excel.py
"""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
from collections import Counter
from pathlib import Path

try:
    import openpyxl
except ImportError as e:  # pragma: no cover
    raise SystemExit("Instale openpyxl: pip install openpyxl") from e

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_XLSM = ROOT / "modelos" / "orçamento" / "ORcAMENTO_OFICIAL_2607281626.xlsm"
OUT_MAPA = ROOT / "backend" / "app" / "data" / "mapa_facas.json"
OUT_CAT = ROOT / "backend" / "app" / "data" / "catalog_oficial.json"

SHEET_MAPA = "MAPA DE FACAS 20260715 ATUAL"

MAQ_MAP = {
    "BETAFLEX": "BETA",
    "BETA": "BETA",
    "ETIRAMA": "ETIRAMA",
    "MODULAR SPX": "MODULAR",
    "MODULAR": "MODULAR",
    "REFLEXO": "160",
    "REFLEXO 160": "160",
    "160": "160",
    "REFLEXO 250": "250",
    "250": "250",
    "BATIDA": "BATIDA",
}


def _norm_spaces(s: str) -> str:
    return " ".join(str(s).strip().split())


def _fmt_num(x, max_decimals: int = 6) -> str:
    if x is None:
        return ""
    if isinstance(x, float):
        if abs(x - round(x)) < 1e-9:
            return str(int(round(x)))
        s = f"{x:.{max_decimals}f}".rstrip("0").rstrip(".")
        return s.replace(".", ",")
    return str(x).replace(".", ",")


def _to_float(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().upper()
    s = s.replace("Ø", "").replace("⌀", "").replace("DIAM", "").replace("Ø", "")
    s = s.replace(",", ".").strip()
    # keep only leading number
    m = re.match(r"^[^\d]*(\d+(?:\.\d+)?)", s)
    if not m:
        try:
            return float(s)
        except ValueError:
            return None
    try:
        return float(m.group(1))
    except ValueError:
        return None


def _normalize_tamanho_text(raw) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s.upper() in ("TAMANHO", "NONE"):
        return None
    # REMALINA etc. — texto livre
    return s.replace(".", ",") if "X" in s.upper() or "Ø" in s.upper() or "⌀" in s else s


def _is_redonda(formato: str | None) -> bool:
    if not formato:
        return False
    return "REDOND" in formato.upper()


def _parse_diametro(tamanho_raw) -> float | None:
    if tamanho_raw is None:
        return None
    return _to_float(tamanho_raw)


def _build_medida(
    formato: str | None,
    tamanho_raw,
    largura: float | None,
    puxada: float | None,
) -> tuple[str | None, str, float | None]:
    """
    Retorna (medida, tipo_tamanho, diametro_cm).
    tipo_tamanho: diametro | retangular | altura | texto | desconhecido
    """
    fmt = (formato or "").strip().upper()
    tam_txt = _normalize_tamanho_text(tamanho_raw)

    if _is_redonda(fmt):
        diam = _parse_diametro(tamanho_raw)
        if diam is not None:
            return f"Ø {_fmt_num(diam)}", "diametro", diam
        if tam_txt:
            # já veio com Ø
            t = tam_txt if "Ø" in tam_txt.upper() or "⌀" in tam_txt else f"Ø {tam_txt}"
            return t.replace("⌀", "Ø"), "diametro", _parse_diametro(tam_txt)
        return None, "diametro", None

    if tam_txt and "X" in tam_txt.upper():
        med = tam_txt.upper().replace(" ", "")
        # normalizar x minúsculo
        med = re.sub(r"[xX]", "X", med)
        return med, "retangular", None

    # RETA / numérico: LARGURA x altura (TAMANHO≈PUXADA na maioria)
    tf = _to_float(tamanho_raw)
    if tf is not None and largura is not None and largura > 0:
        return f"{_fmt_num(largura)}X{_fmt_num(tf)}", "altura", None
    if tf is not None and puxada is not None:
        # sem largura — usar puxada como referência de altura
        return f"{_fmt_num(tf)}", "altura", None
    if tam_txt:
        return tam_txt, "texto", None
    if largura is not None and puxada is not None:
        return f"{_fmt_num(largura)}X{_fmt_num(puxada)}", "composto", None
    return None, "desconhecido", None


def extract_mapa(ws) -> list[dict]:
    facas: list[dict] = []
    skipped: Counter = Counter()

    for r in range(2, ws.max_row + 1):
        maq_raw = ws.cell(r, 1).value
        conjugada = ws.cell(r, 2).value
        fornecedor = ws.cell(r, 3).value
        n_facas = ws.cell(r, 4).value
        z = ws.cell(r, 5).value
        formato_raw = ws.cell(r, 6).value
        tamanho = ws.cell(r, 7).value
        puxada = ws.cell(r, 8).value
        largura = ws.cell(r, 9).value
        rep = ws.cell(r, 10).value
        cil = ws.cell(r, 11).value
        col = ws.cell(r, 12).value
        cliente = ws.cell(r, 13).value

        if maq_raw is None:
            skipped["sem_maquina"] += 1
            continue
        if str(maq_raw).strip().upper() in ("MAQUINA", "MÁQUINA"):
            skipped["header"] += 1
            continue
        if str(fornecedor or "").strip().upper() == "FORNECEDOR":
            skipped["header_row"] += 1
            continue
        if str(formato_raw or "").strip().upper() == "FORMATO":
            skipped["header_fmt"] += 1
            continue

        # linhas separadoras só com máquina
        formato = None if formato_raw is None else _norm_spaces(str(formato_raw))
        if not formato and tamanho is None and puxada is None and z is None:
            skipped["separador"] += 1
            continue

        maq_u = _norm_spaces(maq_raw).upper()
        if maq_u not in MAQ_MAP:
            skipped[f"maq_desconhecida:{maq_raw}"] += 1
            continue
        maq_std = MAQ_MAP[maq_u]

        z_f = _to_float(z)
        pux_f = _to_float(puxada)
        larg_f = _to_float(largura)
        # REP = REPETIÇÃO (pode ser fração manual / calculada)
        rep_f = _to_float(rep) if not (
            isinstance(rep, str) and set(str(rep).strip()) <= {"*", "-"}
        ) else None
        if isinstance(rep, str) and rep_f is None and str(rep).strip() not in ("", "*", "***", "-"):
            # texto estranho em REP
            rep_f = None

        medida, tipo_tam, diametro = _build_medida(formato, tamanho, larg_f, pux_f)
        if not medida and not formato:
            skipped["sem_conteudo"] += 1
            continue

        incompleta = pux_f is None or pux_f <= 0 or z_f is None
        # ainda assim listar se tiver medida/formato (planilha analógica)
        if incompleta and not medida and not formato:
            skipped["incompleta_vazia"] += 1
            continue

        n_f = None
        if n_facas is not None:
            try:
                n_f = int(float(n_facas))
            except (TypeError, ValueError):
                n_f = None

        nota = None if cliente is None else (str(cliente).strip() or None)
        conj = None if conjugada is None else (_norm_spaces(str(conjugada)) or None)
        forn = None if fornecedor is None else (_norm_spaces(str(fornecedor)) or None)

        # Label legível para o seletor
        parts = []
        if medida:
            parts.append(medida)
        if formato:
            parts.append(formato)
        parts.append(maq_std)
        if z_f is not None:
            parts.append(f"Z={_fmt_num(z_f)}")
        if rep_f is not None:
            parts.append(f"REP={_fmt_num(rep_f)}")
        if pux_f is not None:
            parts.append(f"pux={_fmt_num(pux_f)}")
        elif incompleta:
            parts.append("pux=MANUAL")
        if _is_redonda(formato):
            parts.append("diâmetro")
        label = " · ".join(parts)

        tamanho_exibicao = None
        if tamanho is not None:
            tamanho_exibicao = str(tamanho).strip()

        facas.append(
            {
                "id": r,
                "medida": medida or "",
                "tamanho_raw": tamanho_exibicao,
                "tamanho_tipo": tipo_tam,
                "diametro_cm": diametro,
                "formato": formato,
                "faca": formato,  # compat: UI antiga usava "faca" = formato
                "puxada": pux_f,
                "z": z_f,
                "repeticao": rep_f,
                "maquina_catalogo": maq_std,
                "maquina_origem": maq_u,
                "largura_faca": larg_f,
                "n_facas": n_f,
                "cilindro": None if cil is None else str(cil).strip() or None,
                "colunas_mapa": None if col is None else str(col).strip() or None,
                "conjugada": conj,
                "fornecedor": forn,
                "cliente_nota": nota,
                "completa": not incompleta,
                "label": label,
            }
        )

    print("Facas:", len(facas), "| skip:", dict(skipped))
    print("Máquinas:", dict(Counter(f["maquina_catalogo"] for f in facas)))
    print("Formatos:", dict(Counter(f.get("formato") or "?" for f in facas)))
    print("Completas:", sum(1 for f in facas if f["completa"]),
          "| incompletas (puxada/Z manual):", sum(1 for f in facas if not f["completa"]))
    print("REDONDA:", sum(1 for f in facas if _is_redonda(f.get("formato"))))
    return facas


def sync_catalog(wb, cat: dict) -> dict:
    beta_rates = {
        "0": 120.0, "1": 145.0, "2": 155.0, "3": 165.0,
        "4": 175.0, "4V": 180.0, "5": 190.0,
    }
    batida_rates = {"1": 200.0, "2": 210.0}
    modular_rates = {
        "0": 120.0, "1": 150.0, "2": 160.0, "3": 170.0,
        "4": 190.0, "4V": 200.0, "5": 210.0, "6": 220.0, "7": 230.0, "8": 240.0,
    }
    cat["hora_maquina"] = {
        "BETA": dict(beta_rates),
        "160": dict(beta_rates),
        "250": dict(beta_rates),
        "ETIRAMA": dict(beta_rates),
        "BATIDA": dict(batida_rates),
        "MODULAR": dict(modular_rates),
    }
    cat["maquinas"] = ["BETA", "160", "250", "ETIRAMA", "BATIDA", "MODULAR"]
    cat["maquinas_roda_servico"] = list(cat["maquinas"])
    cat["maquina_aliases"] = {
        "BETA / 160  / 250 / ETIRAMA": "BETA",
        "BETA / 160 / 250 / ETIRAMA": "BETA",
        "BETAFLEX": "BETA",
        "REFLEXO": "160",
        "REFLEXO 160": "160",
        "REFLEXO 250": "250",
        "MODULAR SPX": "MODULAR",
    }
    cat["maquina_origem_mapa"] = {
        "BETAFLEX": "BETA",
        "ETIRAMA": "ETIRAMA",
        "MODULAR SPX": "MODULAR",
        "REFLEXO": "160",
        "REFLEXO 160": "160",
        "REFLEXO 250": "250",
        "BATIDA": "BATIDA",
    }

    # --- MEDIDA_CAIXAS (fonte) + CAIXAS (tabela gerada) ---
    # Empacotamento profissional: qtde = CEILING(rolos / rolos_por_caixa)
    # Capacidades alinhadas aos trechos sãos da sheet CAIXAS legada.
    empacotamento = {
        '1"': {"caixa_id": 2, "medida": "250x200x200", "rolos_por_caixa": 20},
        '3"': {"caixa_id": 6, "medida": "500x300x300", "rolos_por_caixa": 12},
    }
    medida_caixas: list[dict] = []
    if "MEDIDA_CAIXAS" in wb.sheetnames:
        ws_m = wb["MEDIDA_CAIXAS"]
        for r in range(2, ws_m.max_row + 1):
            cid, med = ws_m.cell(r, 1).value, ws_m.cell(r, 2).value
            if cid is None or med is None:
                continue
            try:
                cid_i = int(float(cid))
            except (TypeError, ValueError):
                continue
            parts = str(med).lower().replace(" ", "").split("x")
            comp = larg = alt = None
            if len(parts) == 3:
                try:
                    comp, larg, alt = (int(float(p)) for p in parts)
                except (TypeError, ValueError):
                    pass
            # Colunas opcionais enriquecidas (C..I) se existirem
            rolos_1 = ws_m.cell(r, 6).value
            rolos_3 = ws_m.cell(r, 7).value
            pref_1 = ws_m.cell(r, 8).value
            pref_3 = ws_m.cell(r, 9).value
            item = {
                "id": cid_i,
                "medida": str(med).strip(),
                "comp_mm": comp,
                "larg_mm": larg,
                "alt_mm": alt,
                "preferida_1": bool(pref_1) if pref_1 is not None else cid_i == empacotamento['1"']["caixa_id"],
                "preferida_3": bool(pref_3) if pref_3 is not None else cid_i == empacotamento['3"']["caixa_id"],
                "rolos_1": int(rolos_1) if isinstance(rolos_1, (int, float)) else (
                    empacotamento['1"']["rolos_por_caixa"] if cid_i == empacotamento['1"']["caixa_id"] else None
                ),
                "rolos_3": int(rolos_3) if isinstance(rolos_3, (int, float)) else (
                    empacotamento['3"']["rolos_por_caixa"] if cid_i == empacotamento['3"']["caixa_id"] else None
                ),
            }
            medida_caixas.append(item)
            # Se a sheet marcar preferida + capacidade, atualiza empacotamento
            if item["preferida_1"] and item["rolos_1"]:
                empacotamento['1"'] = {
                    "caixa_id": cid_i,
                    "medida": item["medida"],
                    "rolos_por_caixa": int(item["rolos_1"]),
                }
            if item["preferida_3"] and item["rolos_3"]:
                empacotamento['3"'] = {
                    "caixa_id": cid_i,
                    "medida": item["medida"],
                    "rolos_por_caixa": int(item["rolos_3"]),
                }
    else:
        # Fallback: medidas do estoque / sheet nova ainda não salva no xlsm
        medida_caixas = [
            {"id": 1, "medida": "200x150x120", "comp_mm": 200, "larg_mm": 150, "alt_mm": 120,
             "preferida_1": False, "preferida_3": False, "rolos_1": None, "rolos_3": None},
            {"id": 2, "medida": "250x200x200", "comp_mm": 250, "larg_mm": 200, "alt_mm": 200,
             "preferida_1": True, "preferida_3": False, "rolos_1": 20, "rolos_3": None},
            {"id": 3, "medida": "250x250x200", "comp_mm": 250, "larg_mm": 250, "alt_mm": 200,
             "preferida_1": False, "preferida_3": False, "rolos_1": None, "rolos_3": None},
            {"id": 4, "medida": "400x300x223", "comp_mm": 400, "larg_mm": 300, "alt_mm": 223,
             "preferida_1": False, "preferida_3": False, "rolos_1": None, "rolos_3": None},
            {"id": 5, "medida": "300x300x317", "comp_mm": 300, "larg_mm": 300, "alt_mm": 317,
             "preferida_1": False, "preferida_3": False, "rolos_1": None, "rolos_3": None},
            {"id": 6, "medida": "500x300x300", "comp_mm": 500, "larg_mm": 300, "alt_mm": 300,
             "preferida_1": False, "preferida_3": True, "rolos_1": None, "rolos_3": 12},
            {"id": 7, "medida": "460x360x340", "comp_mm": 460, "larg_mm": 360, "alt_mm": 340,
             "preferida_1": False, "preferida_3": False, "rolos_1": None, "rolos_3": None},
            {"id": 8, "medida": "500x400x300", "comp_mm": 500, "larg_mm": 400, "alt_mm": 300,
             "preferida_1": False, "preferida_3": False, "rolos_1": None, "rolos_3": None},
            {"id": 9, "medida": "540x405x335", "comp_mm": 540, "larg_mm": 405, "alt_mm": 335,
             "preferida_1": False, "preferida_3": False, "rolos_1": None, "rolos_3": None},
        ]

    def _gen_caixas(emp: dict, max_3: int = 500, max_1: int = 2000) -> dict[str, int]:
        out: dict[str, int] = {}
        c3 = int(emp['3"']["rolos_por_caixa"])
        c1 = int(emp['1"']["rolos_por_caixa"])
        for n in range(1, max_3 + 1):
            out[f'3"{n}'] = int(math.ceil(n / c3))
        for n in range(1, max_1 + 1):
            out[f'1"{n}'] = int(math.ceil(n / c1))
        return out

    cat["medida_caixas"] = medida_caixas
    cat["caixa_empacotamento"] = empacotamento
    cat["caixas"] = _gen_caixas(empacotamento)
    cat["preco_caixa"] = float(cat.get("preco_caixa", 7))

    ws_t = wb["TINTA"]
    cat["tinta"] = {
        "faixa_m2": float(ws_t["B2"].value or 30),
        "valor_ate_30_por_cor": float(ws_t["C2"].value or 10),
        "valor_acima_m2": float(ws_t["C3"].value or 0.4),
    }

    ws_a = wb["ACABAMENTOS"]
    acab = {}
    for r in range(2, ws_a.max_row + 1):
        nome, val = ws_a.cell(r, 1).value, ws_a.cell(r, 2).value
        if not nome or val is None:
            continue
        if isinstance(val, str) and val.startswith("="):
            val = 3.5
        acab[_norm_spaces(nome)] = float(val)
    cat["acabamentos"] = acab

    ws_pa = wb["PERDA DE ACABAMENTO"]
    perda_ac = {}
    for r in range(2, ws_pa.max_row + 1):
        nome, val = ws_pa.cell(r, 1).value, ws_pa.cell(r, 2).value
        if not nome or val is None:
            continue
        perda_ac[_norm_spaces(nome)] = float(val)
    cat["perda_acabamento"] = perda_ac

    ws_p = wb["PAPEL"]
    papel = {}
    for r in range(2, ws_p.max_row + 1):
        nome, val = ws_p.cell(r, 1).value, ws_p.cell(r, 2).value
        if not nome or val is None:
            continue
        papel[_norm_spaces(nome)] = float(val)
    cat["papel"] = papel

    ws_ppa = wb["PERDA DE PAPEL ACERTO"]
    ppa = {}
    for r in range(2, ws_ppa.max_row + 1):
        c, ml = ws_ppa.cell(r, 1).value, ws_ppa.cell(r, 3).value
        if c is None:
            continue
        if ml is None:
            ml = 0
        key = (
            str(int(c))
            if isinstance(c, (int, float)) and c == int(c)
            else str(c).strip()
        )
        ppa[key] = float(ml)
    if "4V" not in ppa:
        ppa["4V"] = float(ppa.get("4", 200))
    cat["perda_papel_acerto_ml"] = ppa
    return cat


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsm", type=Path, default=DEFAULT_XLSM)
    ap.add_argument("--skip-catalog", action="store_true")
    args = ap.parse_args(argv)
    if not args.xlsm.exists():
        print("Arquivo não encontrado:", args.xlsm, file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(args.xlsm, data_only=True, keep_vba=True)
    if SHEET_MAPA not in wb.sheetnames:
        print("Sheet ausente:", SHEET_MAPA, file=sys.stderr)
        return 1

    facas = extract_mapa(wb[SHEET_MAPA])
    OUT_MAPA.write_text(json.dumps(facas, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print("Wrote", OUT_MAPA)

    if not args.skip_catalog:
        cat = json.loads(OUT_CAT.read_text(encoding="utf-8")) if OUT_CAT.exists() else {}
        cat = sync_catalog(wb, cat)
        OUT_CAT.write_text(json.dumps(cat, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        emp = cat.get("caixa_empacotamento", {})
        print(
            "Wrote",
            OUT_CAT,
            "| caixas=",
            len(cat["caixas"]),
            "| medida_caixas=",
            len(cat.get("medida_caixas") or []),
            "| emp=",
            {k: v.get("rolos_por_caixa") for k, v in emp.items()},
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
