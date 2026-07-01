from __future__ import annotations

from collections import defaultdict
from decimal import Decimal

from openpyxl import load_workbook

from app.domain.b3_schema import (
    B3Column,
    CUSTODY_TRANSFER_MOVEMENTS,
    DIRECTION_CREDIT,
    DIRECTION_DEBIT,
    INCOME_VALUE_QUANTIZE,
    MIN_COLUMNS,
    REQUIRED_HEADERS,
)
from app.domain.models import Movement
from app.domain.movement_classifier import classify_movement
from app.domain.position_calculator import PositionCalculator
from app.domain.transfer_resolver import TransferKey, classify_transfer
from app.domain.value_objects import BrazilianDate, Money, MovementKind, Ticker


class B3MovementParser:
    """Parser do XLSX oficial de movimentação da Área do Investidor B3."""

    def parse_file(self, file_path: str) -> list[Movement]:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        if not self._validate_headers(rows[0]):
            workbook.close()
            raise ValueError("Layout do arquivo não corresponde ao XLSX oficial da B3.")

        custody_keys = self._custody_transfer_keys_from_rows(rows[1:])

        movements: list[Movement] = []
        for row in rows[1:]:
            movement = self._parse_row(row, custody_keys)
            if movement:
                movements.append(movement)

        workbook.close()
        return movements

    @staticmethod
    def _custody_transfer_keys_from_rows(rows: list[tuple]) -> set[TransferKey]:
        directions: dict[TransferKey, set[str]] = defaultdict(set)
        for row in rows:
            if not row or len(row) < MIN_COLUMNS:
                continue
            movement_label = str(row[B3Column.MOVEMENT] or "").strip()
            if movement_label not in CUSTODY_TRANSFER_MOVEMENTS:
                continue
            trade_date = BrazilianDate.parse(row[B3Column.TRADE_DATE])
            product = str(row[B3Column.PRODUCT] or "").strip()
            if not trade_date or not product:
                continue
            key = TransferKey(
                trade_date.value,
                product,
                B3MovementParser._to_decimal(row[B3Column.QUANTITY]),
            )
            directions[key].add(str(row[B3Column.DIRECTION] or "").strip().lower())

        return {
            key
            for key, seen in directions.items()
            if DIRECTION_CREDIT in seen and DIRECTION_DEBIT in seen
        }

    @staticmethod
    def _validate_headers(header_row: tuple) -> bool:
        if not header_row or len(header_row) < len(REQUIRED_HEADERS):
            return False
        normalized = [str(cell or "").strip().lower() for cell in header_row]
        return all(normalized[idx] == expected for idx, expected in enumerate(REQUIRED_HEADERS))

    def _parse_row(self, row: tuple, custody_keys: set[TransferKey]) -> Movement | None:
        if not row or len(row) < MIN_COLUMNS:
            return None

        direction = str(row[B3Column.DIRECTION] or "").strip()
        trade_date = BrazilianDate.parse(row[B3Column.TRADE_DATE])
        movement_label = str(row[B3Column.MOVEMENT] or "").strip()
        product = str(row[B3Column.PRODUCT] or "").strip()
        institution = str(row[B3Column.INSTITUTION] or "").strip()

        if not trade_date or not movement_label or not product:
            return None

        ticker_vo = Ticker.from_product(product)
        if not ticker_vo:
            return None

        kind, income_type = classify_movement(movement_label, direction)
        if movement_label in CUSTODY_TRANSFER_MOVEMENTS:
            kind = classify_transfer(
                trade_date=trade_date.value,
                product=product,
                quantity=self._to_decimal(row[B3Column.QUANTITY]),
                direction=direction,
                custody_keys=custody_keys,
            )
            if kind == MovementKind.IGNORE:
                return None
            income_type = None
        elif kind.value == "IGNORE":
            return None

        quantity = self._to_decimal(row[B3Column.QUANTITY])
        unit_price = Money.parse_br(row[B3Column.UNIT_PRICE])
        total_value = Money.parse_br(row[B3Column.TOTAL_VALUE])

        if kind.value == "INCOME" and total_value is None and unit_price and quantity:
            total_value = Money(
                amount=(unit_price.amount * quantity).quantize(Decimal(INCOME_VALUE_QUANTIZE))
            )

        external_key = PositionCalculator.external_key(
            trade_date.value.isoformat(),
            movement_label,
            product,
            direction,
            str(quantity),
            str(total_value.amount if total_value else ""),
        )

        return Movement(
            trade_date=trade_date.value,
            kind=kind,
            ticker=ticker_vo.code,
            product=product,
            institution=institution,
            direction=direction,
            movement_label=movement_label,
            quantity=quantity,
            unit_price=unit_price.amount if unit_price else None,
            total_value=total_value.amount if total_value else None,
            income_type=income_type,
            external_key=external_key,
        )

    @staticmethod
    def _to_decimal(value: object) -> Decimal:
        if value is None or value == "-" or value == "":
            return Decimal("0")
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        text = str(value).strip().replace(".", "").replace(",", ".")
        if not text:
            return Decimal("0")
        return Decimal(text)

    @staticmethod
    def company_name(product: str) -> str:
        parts = product.split("-", 1)
        if len(parts) == 2:
            return parts[1].strip()
        return product
