#!/usr/bin/env python3
"""
Sincroniza catalog_oficial.json (+ mapa_facas.json) a partir do ORcAMENTO_OFICIAL rv4.

Fonte canônica: sheet ORÇAMENTO e abas referenciadas por ela (rv4).
Mapa de facas: sheet 'MAPA DE FACAS 20260715 ATUAL' (pivot MAPA_DE_FACAS → ORÇAMENTO).

Uso:
  python3 scripts/sync_orcamento_from_xlsm.py \\
    --xlsm /caminho/ORcAMENTO_OFICIAL_rv4.xlsm
  python3 scripts/sync_orcamento_from_xlsm.py --skip-mapa
"""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
from collections import Counter
from pathlib import Path

try:
    import openpyxl
except ImportError as e:  # pragma: no cover
    raise SystemExit("Instale openpyxl: pip install openpyxl") from e

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSM = ROOT / "apps/api/resources/data/orcamento/ORcAMENTO_OFICIAL_rv4.xlsm"
if not DEFAULT_XLSM.exists():
    DEFAULT_XLSM = Path("/home/dfmoura/Downloads/ORcAMENTO_OFICIAL_rv4.xlsm")
OUT_CAT = ROOT / "apps/api/resources/data/orcamento/catalog_oficial.json"
OUT_MAPA = ROOT / "apps/api/resources/data/orcamento/mapa_facas.json"

SHEET_MAPA = "MAPA DE FACAS 20260715 ATUAL"
SHEET_PERDA_MAQ = "PERDA DE ACERTO "
SHEET_PERDA_TROCA = "PERDA DE ACERTO"

MAQ_MAP = {
    "BETAFLEX": "BETA",
    "BETA": "BETA",
    "ETIRAMA": "ETIRAMA",
    "MODULAR SPX": "MODULAR",
    "MODULAR": "MODULAR",
    "REFLEXO": "160",
    "REFLEXO 160": "160",
    "160": "160",
    "REFLEXO 250": "250",
    "250": "250",
    "BATIDA": "BATIDA",
}


def _norm_spaces(s: str) -> str:
    return " ".join(str(s).strip().split())


def _fmt_num(x, max_decimals: int = 6) -> str:
    if x is None:
        return ""
    if isinstance(x, (int, float)):
        if abs(x - round(x)) < 1e-9:
            return str(int(round(x)))
        s = f"{x:.{max_decimals}f}".rstrip("0").rstrip(".")
        return s.replace(".", ",")
    return str(x).replace(".", ",")


def _to_float(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().upper()
    s = s.replace("Ø", "").replace("⌀", "").replace("DIAM", "")
    s = s.replace(",", ".").strip()
    m = re.match(r"^[^\d]*(\d+(?:\.\d+)?)", s)
    if not m:
        try:
            return float(s)
        except ValueError:
            return None
    try:
        return float(m.group(1))
    except ValueError:
        return None


def _cores_key(raw) -> str:
    if raw is None:
        return ""
    if isinstance(raw, str):
        return raw.strip().upper()
    if isinstance(raw, float) and raw == int(raw):
        return str(int(raw))
    return str(raw).strip()


def _normalize_tamanho_text(raw) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s.upper() in ("TAMANHO", "NONE"):
        return None
    return s.replace(".", ",") if "X" in s.upper() or "Ø" in s.upper() or "⌀" in s else s


def _is_redonda(formato: str | None) -> bool:
    return bool(formato and "REDOND" in formato.upper())


def _parse_diametro(tamanho_raw) -> float | None:
    return _to_float(tamanho_raw)


def _build_medida(formato, tamanho_raw, largura, puxada):
    fmt = (formato or "").strip().upper()
    tam_txt = _normalize_tamanho_text(tamanho_raw)
    if _is_redonda(fmt):
        diam = _parse_diametro(tamanho_raw)
        if diam is not None:
            return f"Ø {_fmt_num(diam)}", "diametro", diam
        if tam_txt:
            t = tam_txt if "Ø" in tam_txt.upper() or "⌀" in tam_txt else f"Ø {tam_txt}"
            return t.replace("⌀", "Ø"), "diametro", _parse_diametro(tam_txt)
        return None, "diametro", None
    if tam_txt and "X" in tam_txt.upper():
        med = re.sub(r"[xX]", "X", tam_txt.upper().replace(" ", ""))
        return med, "retangular", None
    tf = _to_float(tamanho_raw)
    if tf is not None and largura is not None and largura > 0:
        return f"{_fmt_num(largura)}X{_fmt_num(tf)}", "altura", None
    if tf is not None and puxada is not None:
        return f"{_fmt_num(tf)}", "altura", None
    if tam_txt:
        return tam_txt, "texto", None
    if largura is not None and puxada is not None:
        return f"{_fmt_num(largura)}X{_fmt_num(puxada)}", "composto", None
    return None, "desconhecido", None


def extract_mapa(ws) -> list[dict]:
    facas: list[dict] = []
    skipped: Counter = Counter()
    for r in range(2, ws.max_row + 1):
        maq_raw = ws.cell(r, 1).value
        conjugada = ws.cell(r, 2).value
        fornecedor = ws.cell(r, 3).value
        n_facas = ws.cell(r, 4).value
        z = ws.cell(r, 5).value
        formato_raw = ws.cell(r, 6).value
        tamanho = ws.cell(r, 7).value
        puxada = ws.cell(r, 8).value
        largura = ws.cell(r, 9).value
        rep = ws.cell(r, 10).value
        cil = ws.cell(r, 11).value
        col = ws.cell(r, 12).value
        cliente = ws.cell(r, 13).value
        if maq_raw is None:
            skipped["sem_maquina"] += 1
            continue
        if str(maq_raw).strip().upper() in ("MAQUINA", "MÁQUINA"):
            skipped["header"] += 1
            continue
        if str(fornecedor or "").strip().upper() == "FORNECEDOR":
            skipped["header_row"] += 1
            continue
        if str(formato_raw or "").strip().upper() == "FORMATO":
            skipped["header_fmt"] += 1
            continue
        formato = None if formato_raw is None else _norm_spaces(str(formato_raw))
        if not formato and tamanho is None and puxada is None and z is None:
            skipped["separador"] += 1
            continue
        maq_u = _norm_spaces(maq_raw).upper()
        if maq_u not in MAQ_MAP:
            skipped[f"maq_desconhecida:{maq_raw}"] += 1
            continue
        maq_std = MAQ_MAP[maq_u]
        z_f = _to_float(z)
        pux_f = _to_float(puxada)
        larg_f = _to_float(largura)
        rep_f = _to_float(rep) if not (
            isinstance(rep, str) and set(str(rep).strip()) <= {"*", "-"}
        ) else None
        medida, tipo_tam, diametro = _build_medida(formato, tamanho, larg_f, pux_f)
        if not medida and not formato:
            skipped["sem_conteudo"] += 1
            continue
        incompleta = pux_f is None or pux_f <= 0 or z_f is None
        if incompleta and not medida and not formato:
            skipped["incompleta_vazia"] += 1
            continue
        n_f = None
        if n_facas is not None:
            try:
                n_f = int(float(n_facas))
            except (TypeError, ValueError):
                n_f = None
        nota = None if cliente is None else (str(cliente).strip() or None)
        conj = None if conjugada is None else (_norm_spaces(str(conjugada)) or None)
        forn = None if fornecedor is None else (_norm_spaces(str(fornecedor)) or None)
        parts = []
        if medida:
            parts.append(medida)
        if formato:
            parts.append(formato)
        parts.append(maq_std)
        if z_f is not None:
            parts.append(f"Z={_fmt_num(z_f)}")
        if rep_f is not None:
            parts.append(f"REP={_fmt_num(rep_f)}")
        if pux_f is not None:
            parts.append(f"pux={_fmt_num(pux_f)}")
        elif incompleta:
            parts.append("pux=MANUAL")
        if _is_redonda(formato):
            parts.append("diâmetro")
        label = " · ".join(parts)
        tamanho_exibicao = None if tamanho is None else str(tamanho).strip()
        facas.append(
            {
                "id": r,
                "medida": medida or "",
                "tamanho_raw": tamanho_exibicao,
                "tamanho_tipo": tipo_tam,
                "diametro_cm": diametro,
                "formato": formato,
                "faca": formato,
                "puxada": pux_f,
                "z": z_f,
                "repeticao": rep_f,
                "maquina_catalogo": maq_std,
                "maquina_origem": maq_u,
                "largura_faca": larg_f,
                "n_facas": n_f,
                "cilindro": None if cil is None else str(cil).strip() or None,
                "colunas_mapa": None if col is None else str(col).strip() or None,
                "conjugada": conj,
                "fornecedor": forn,
                "cliente_nota": nota,
                "completa": not incompleta,
                "label": label,
            }
        )
    print("Facas:", len(facas), "| skip:", dict(skipped))
    return facas


def _read_hora_maquina(ws) -> dict[str, dict[str, float]]:
    """Lê blocos BETA/BATIDA/MODULAR da sheet HORA MÁQUINA."""
    blocks = [
        ("BETA", 3, 5),
        ("160", 3, 5),
        ("250", 3, 5),
        ("ETIRAMA", 3, 5),
        ("BATIDA", 10, 12),
        ("MODULAR", 17, 19),
    ]
    out: dict[str, dict[str, float]] = {}
    for maq, col_cores, col_rate in blocks:
        rates: dict[str, float] = {}
        for r in range(2, ws.max_row + 1):
            cores = ws.cell(r, col_cores).value
            rate = ws.cell(r, col_rate).value
            if cores is None or rate is None:
                continue
            key = _cores_key(cores)
            if key == "":
                continue
            rates[key] = float(rate)
        if rates:
            out[maq] = rates
    # BETA/160/250/ETIRAMA compartilham a mesma tabela na coluna B
    shared = out.get("BETA", {})
    for alias in ("160", "250", "ETIRAMA"):
        if alias not in out and shared:
            out[alias] = dict(shared)
    return out


def _read_perda_troca(ws) -> dict[str, float]:
    """PERDA DE ACERTO — col E (m² fator @ largura ref) × larg/100 × modelos (ORÇAMENTO I14+)."""
    out: dict[str, float] = {}
    for r in range(2, ws.max_row + 1):
        cores = ws.cell(r, 1).value
        fator = ws.cell(r, 5).value
        if cores is None:
            continue
        key = _cores_key(cores)
        out[key] = 0.0 if fator is None else float(fator)
    return out


def _read_perda_maquina(ws) -> tuple[dict[str, float], dict[str, float]]:
    """PERDA DE ACERTO (espaço) — m² fixos 0–3 (col D) e ml 4+ (col B)."""
    fixed: dict[str, float] = {}
    ml_4plus: dict[str, float] = {}
    for r in range(2, ws.max_row + 1):
        cores = ws.cell(r, 1).value
        ml = ws.cell(r, 2).value
        m2 = ws.cell(r, 4).value
        if cores is None:
            continue
        key = _cores_key(cores)
        if key in ("0", "1", "2", "3") and m2 is not None:
            fixed[key] = float(m2)
        elif ml is not None:
            ml_4plus[key] = float(ml)
    return fixed, ml_4plus


def _read_tinta_matriz(ws) -> dict:
    thresholds: list[float] = []
    rates: dict[str, list[float]] = {"1": [], "2": [], "3": [], "4": []}
    for r in range(2, ws.max_row + 1):
        th = ws.cell(r, 1).value
        if th is None:
            continue
        thresholds.append(float(th))
        for i, col_key in enumerate(["1", "2", "3", "4"], start=2):
            val = ws.cell(r, i).value
            rates[col_key].append(float(val) if val is not None else 0.0)
    return {"thresholds": thresholds, "rates": rates}


def _read_caixas(ws) -> dict[str, int]:
    out: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        chave = ws.cell(r, 2).value
        qtde = ws.cell(r, 5).value
        if chave is None or qtde is None:
            continue
        out[str(chave).strip()] = int(float(qtde))
    return out


def sync_catalog(wb, cat: dict) -> dict:
    cat["fonte_xlsm"] = "ORcAMENTO_OFICIAL_rv4"
    cat["motor_version_catalogo"] = 2

    cat["hora_maquina"] = _read_hora_maquina(wb["HORA MÁQUINA "])
    cat["maquinas"] = ["BETA", "160", "250", "ETIRAMA", "BATIDA", "MODULAR"]
    cat["maquinas_roda_servico"] = list(cat["maquinas"])
    cat["maquina_aliases"] = {
        "BETA / 160  / 250 / ETIRAMA": "BETA",
        "BETA / 160 / 250 / ETIRAMA": "BETA",
        "BETAFLEX": "BETA",
        "REFLEXO": "160",
        "REFLEXO 160": "160",
        "REFLEXO 250": "250",
        "MODULAR SPX": "MODULAR",
    }
    cat["maquina_origem_mapa"] = {
        "BETAFLEX": "BETA",
        "ETIRAMA": "ETIRAMA",
        "MODULAR SPX": "MODULAR",
        "REFLEXO": "160",
        "REFLEXO 160": "160",
        "REFLEXO 250": "250",
        "BATIDA": "BATIDA",
    }

    ws_t = wb["TINTA"]
    cat["tinta"] = {
        "faixa_m2": float(ws_t["B2"].value or 30),
        "valor_ate_30_por_cor": float(ws_t["C2"].value or 10),
        "valor_acima_m2": float(ws_t["C3"].value or 0.4),
    }
    if "TINTA (2)" in wb.sheetnames:
        cat["tinta_matriz"] = _read_tinta_matriz(wb["TINTA (2)"])

    ws_a = wb["ACABAMENTOS"]
    acab = {}
    for r in range(2, ws_a.max_row + 1):
        nome, val = ws_a.cell(r, 1).value, ws_a.cell(r, 2).value
        if not nome or val is None:
            continue
        if isinstance(val, str) and val.startswith("="):
            val = 3.5
        acab[_norm_spaces(nome)] = float(val)
    cat["acabamentos"] = acab

    ws_pa = wb["PERDA DE ACABAMENTO"]
    perda_ac = {}
    for r in range(2, ws_pa.max_row + 1):
        nome, val = ws_pa.cell(r, 1).value, ws_pa.cell(r, 2).value
        if not nome or val is None:
            continue
        perda_ac[_norm_spaces(nome)] = float(val)
    cat["perda_acabamento"] = perda_ac

    ws_p = wb["PAPEL"]
    papel = {}
    for r in range(2, ws_p.max_row + 1):
        nome, val = ws_p.cell(r, 1).value, ws_p.cell(r, 2).value
        if not nome or val is None:
            continue
        papel[_norm_spaces(nome)] = float(val)
    cat["papel"] = papel

    perda03, ml4 = _read_perda_maquina(wb[SHEET_PERDA_MAQ])
    cat["perda_papel_0_3"] = perda03
    # ORÇAMENTO J14+ usa metros fixos 4V–8 (não col B da aba de perda).
    cat["perda_papel_f6"] = float(ml4.get("4", 150.0))
    cat["perda_acerto_m_4v"] = 250.0
    cat["perda_acerto_m_5"] = 250.0
    cat["perda_acerto_m_6"] = 260.0
    cat["perda_acerto_m_7"] = 270.0
    cat["perda_acerto_m_8"] = 280.0

    cat["perda_troca_m2_fator"] = _read_perda_troca(wb[SHEET_PERDA_TROCA])
    # Compat legado (ml ≈ fator / 0.085 ref rv4)
    cat["perda_papel_acerto_ml"] = {
        k: (0.0 if v == 0 else round(v / 0.085, 6))
        for k, v in cat["perda_troca_m2_fator"].items()
    }

    ws_hp = wb["HORA PARADA"]
    hora_parada = {}
    for r in range(2, ws_hp.max_row + 1):
        nome, val = ws_hp.cell(r, 1).value, ws_hp.cell(r, 2).value
        if not nome or val is None:
            continue
        hora_parada[_norm_spaces(nome)] = float(val)
    cat["hora_parada_h"] = hora_parada

    ws_mat = wb["MATRIZ "]
    cat["matriz_cm2"] = float(ws_mat["D3"].value or 0.28)

    ws_tub = wb["TUBETE"]
    tubete = {}
    for r in range(2, ws_tub.max_row + 1):
        tam, val = ws_tub.cell(r, 2).value, ws_tub.cell(r, 3).value
        if not tam or val is None:
            continue
        tubete[str(tam).strip()] = float(val)
    cat["tubete"] = tubete

    ws_cx = wb["CAIXAS"]
    cat["caixas"] = _read_caixas(ws_cx)
    # Metadados empacotamento (rolos por caixa inferidos da tabela)
    cat["caixa_empacotamento"] = {
        '1"': {"caixa_id": 2, "medida": "250x200x200", "rolos_por_caixa": 20},
        '3"': {"caixa_id": 6, "medida": "500x300x300", "rolos_por_caixa": 12},
    }
    for chave, meta in cat["caixa_empacotamento"].items():
        sample = cat["caixas"].get(f'{chave}12')
        if sample:
            meta["rolos_por_caixa"] = 12 if sample == 1 else meta["rolos_por_caixa"]

    cat.setdefault("setup_horas", 1)
    cat.setdefault("limite_metragem_bobina", 1000)
    cat.setdefault("minutos_troca_bobina", 5)
    cat.setdefault("ceiling_etiqueta", 10)
    cat.setdefault("preco_caixa", 7)

    if "medida_caixas" not in cat:
        cat["medida_caixas"] = []

    return cat


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsm", type=Path, default=DEFAULT_XLSM)
    ap.add_argument("--skip-mapa", action="store_true")
    ap.add_argument("--skip-catalog", action="store_true")
    args = ap.parse_args(argv)

    if not args.xlsm.exists():
        print("Arquivo não encontrado:", args.xlsm, file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(args.xlsm, data_only=True, keep_vba=True)

    if not args.skip_mapa:
        if SHEET_MAPA not in wb.sheetnames:
            print("Sheet ausente:", SHEET_MAPA, file=sys.stderr)
            return 1
        facas = extract_mapa(wb[SHEET_MAPA])
        OUT_MAPA.parent.mkdir(parents=True, exist_ok=True)
        OUT_MAPA.write_text(json.dumps(facas, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        print("Wrote", OUT_MAPA, f"({len(facas)} facas)")

    if not args.skip_catalog:
        cat = json.loads(OUT_CAT.read_text(encoding="utf-8")) if OUT_CAT.exists() else {}
        cat = sync_catalog(wb, cat)
        OUT_CAT.parent.mkdir(parents=True, exist_ok=True)
        OUT_CAT.write_text(json.dumps(cat, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        print(
            "Wrote",
            OUT_CAT,
            "| papeis=",
            len(cat.get("papel", {})),
            "| tinta_matriz=",
            len(cat.get("tinta_matriz", {}).get("thresholds", [])),
            "| caixas=",
            len(cat.get("caixas", {})),
        )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
